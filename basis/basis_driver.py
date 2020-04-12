#!/usr/bin/env python3
# coding:utf8
import csv
import random
import traceback
import unittest
import warnings
from enum import unique, Enum
import time
import os
import pymssql
import pymysql
import win32gui
import win32con
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver import FirefoxProfile, ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from selenium.webdriver.support import expected_conditions as EC

'''使用装饰器启动浏览器Start'''
@unique
class BasisBrowser(Enum):
    Chrome = 0
    Firefox = 1
    Edge = 2

class BasisDriver(object):
    '''a simple demo of selenium framework tool'''
    driver = None
    by_split = None
    PATH = os.path.abspath(os.path.dirname(os.getcwd()))
    def __init__(self,browser_type=0,by_split=",",profile=None):
        '''构造方法：实例化BoxDriver的时候用
        ：param char:分隔符
        ：param Brower_profile:
        可选择的参数，如果不传递则是None
        如果传递一个profile,就会按照预先的设定Brower
        去掉遮挡元素的提示框等
        '''
        self.by_split=by_split
        if browser_type == BasisBrowser.Chrome:
            profile = webdriver.ChromeOptions()
            profile.add_experimental_option("excludeSwitches",["ignore-certificate-errors"])
            driver = webdriver.Chrome(chrome_options=profile)
        elif browser_type == BasisBrowser.Firefox:
            if profile is not None:
                profile = FirefoxProfile(profile)
            driver = webdriver.Firefox(firefox_profile=profile)
        elif browser_type == BasisBrowser.Edge:
            driver = webdriver.Edge()
        else:
            driver = webdriver.PhantomJS()
        try:
            self.driver = driver
            self.by_split = by_split
        except Exception:
            print("浏览器驱动版本或位置不正确：如果使用Anaconda Navigator 存放Python安装包，"
                  "Web测试时要把  **driver.exe 的驱动到  **\Anaconda3\Scripts  的目录下！！！")
            raise NameError("Brower %s Not Found!" % browser_type)

#简单方法
# class BasisDriver(object):
#         csv_file = None
#         sql_file = None
#         mysql_connect = None
#         mysql_cursor = None
#         by_split = None
#
#         def __init__(self, browser, by_split=','):
#             self.by_split = by_split
#             if browser == 'Firefox':
#                 driver = webdriver.Firefox()  # 切记勿漏括号
#                 try:
#                     self.driver = driver
#                     self.by_split = by_split
#                 except Exception:  # try部分代码执行有问题时，就会执行except内部的语句
#                     raise NameError('Firefox Not Found!')
#
#             elif browser == 'Chrome':
#                 driver = webdriver.Chrome()
#                 try:
#                     self.driver = driver
#                     self.by_split = by_split
#                 except Exception:
#                     raise NameError('Chrome Not Found!')
#
#             elif browser == 'Edge':
#                 driver = webdriver.Edge()
#                 try:
#                     self.driver = driver
#                     self.by_split = by_split
#                 except Exception:
#                     raise NameError('Edge Not Found!')
#             else:
#                 driver = webdriver.PhantomJS()
#                 # print('Not Found Browser!')
#                 try:
#                     self.driver = driver
#                     self.by_split = by_split
#                 except Exception:
#                     raise NameError("Browser %s Not Found!" % browser)   "#

    def locate_element(self, selector):
        '''定位单个元素'''
        if self.by_split not in selector:  # 例：'i,password'
            return self.driver.find_element_by_id(selector)
        selector_by = selector.split(self.by_split)[0].strip()  # 获取定位方式（split：间隔）
        selector_value = selector.split(self.by_split)[1].strip()  # 获取定位方式对应的值

        if selector_by == 'i' or selector_by == 'id':
            element = self.driver.find_element_by_id(selector_value)
        elif selector_by == 'n' or selector_by == 'name':
            element = self.driver.find_element_by_name(selector_value)
        elif selector_by == 'c' or selector_by == 'class_name':
            element = self.driver.find_element_by_class_name(selector_value)
        elif selector_by == 'l' or selector_by == 'link_text':
            element = self.driver.find_element_by_link_text(u'%s' % selector_value)
        elif selector_by == 'p' or selector_by == 'partial_link_text':
            element = self.driver.find_element_by_partial_link_text(u'%s' % selector_value)
        elif selector_by == 's' or selector_by == 'css_selector':
            element = self.driver.find_element_by_css_selector(selector_value)
        elif selector_by == 'x' or selector_by == 'xpath':
            element = self.driver.find_element_by_xpath(selector_value)
        elif selector_by == 't' or selector_by == 'tag_name':
            element = self.driver.find_element_by_tag_name(selector_value)
        else:
            raise NameError('please enter a valid type of targeting element')

        return element

    def locate_elements(self, selector):
        '''定位一组元素'''
        if self.by_split not in selector:  # 例：'i,password'
            return self.driver.find_elements_by_id(selector)
        selector_by = selector.split(self.by_split)[0].strip()  # 获取定位方式（split：间隔）
        selector_value = selector.split(self.by_split)[1].strip()  # 获取定位方式对应的值

        if selector_by == 'i' or selector_by == 'id':
            elements = self.driver.find_elements_by_id(selector_value)
        elif selector_by == 'n' or selector_by == 'name':
            elements = self.driver.find_elements_by_name(selector_value)
        elif selector_by == 'c' or selector_by == 'class_name':
            elements = self.driver.find_elements_by_class_name(selector_value)
        elif selector_by == 'l' or selector_by == 'link_text':
            elements = self.driver.find_elements_by_link_text(selector_value)
        elif selector_by == 'p' or selector_by == 'partial_link_text':
            elements = self.driver.find_elements_by_partial_link_text(selector_value)
        elif selector_by == 's' or selector_by == 'css_selector':
            elements = self.driver.find_elements_by_css_selector(selector_value)
        elif selector_by == 'x' or selector_by == 'xpath':
            elements = self.driver.find_elements_by_xpath(selector_value)
        elif selector_by == 't' or selector_by == 'tag_name':
            elements = self.driver.find_elements_by_tag_name(selector_value)
        else:
            raise NameError('please enter a valid type of targeting element')

        return elements

    '''抛出断言方法'''

    def assertion(self, first, second, message):

        try:
            unittest.TestCase().assertIn(first, second, message)
            # raise Exception
        except Exception as msg:
            print('msg:',msg)


    def assertion_log(self, first, second, message, photo_path, log_path):
        '''抛出断言生成日志'''
        try:
            unittest.TestCase().assertIn(first, second, message)
        except Exception as msg:
            print(msg)
            self.get_screenshot(self.PATH+photo_path)
            file = open(self.PATH+log_path,"a+")
            traceback.print_exc(file=file)
            file.flush()
            file.close()

    def alert_exist_accept(self):
        '''
        判断alert是否存在，存在则接受
        :return:
        '''
        if EC.alert_is_present:

            print("Alert exists")
            alert = self.driver.switch_to_alert()
            # print(alert.text)
            alert.accept()
            print("Alert accepted")
        else:

            print("NO alert exists")
    def alert_dismiss(self):
        '''取消alert框'''
        self.driver.switch_to.alert().dismiss()
    def add_cookies(self, cookies):
        '''clear all cookies after driver init'''
        self.driver.delete_all_cookies()

    def add_cookie(self, cookie_dict):
        '''
        Add single cookie by dict
        添加 单个 cookie
        如果该 cookie 已经存在，先删除，再添加
        :param cookie_dict: 字典类型，有两个 key:name 和 value
        :return: 
        '''
        cookie_name = cookie_dict["name"]
        cookie_value = self.driver.get_cookie(cookie_name)
        if cookie_value is not None:
            self.driver.delete_cookie(cookie_name)
        self.driver.add_cookie(cookie_dict)

    def clear_cookies(self):
        '''clear all cookies after driver init'''
        self.driver.delete_all_cookies()

    def close_csv_file(self):
        self.csv_file.close()

    def click(self, selector):
        '''鼠标左键单击选择器 
        To click the selector by the left button of mouse.
        :param selector: 
        :return: 
        '''
        ele = self.locate_element(selector)
        ele.click()

    def click_double(self, selector):
        '''
        鼠标左键双击选择器
        :param selector: 
        :return: 
        '''
        doubleClick = self.locate_element(selector)
        ActionChains(self.driver).double_click(doubleClick).perform()

    def click_right(self, selector):
        '''鼠标右键单击选择器 
        To click the selector by the right button of mouse.
        :param selector: 
        :return: 
        '''
        ele = self.locate_element(selector)
        ac = ActionChains(self.driver)
        ac.context_click(ele).perform()

    def click_by_enter(self, selector):
        '''
        It can type any text / image can be located with ENTER key
        Usage: 
        driver.click_by_ener("i,ele"): 
        '''
        ele = self.locate_element(selector)
        ele.send_keys(Keys.ENTER)

    def close_browser(self):
        '''关闭当前浏览器窗口'''
        self.driver.close()

    def current_window_handle(self):
        '''获取当前窗口句柄'''
        return self.driver.current_window_handle
    def all_window_handles(self):
        '''获取多有窗口句柄'''
        return self.driver.window_handles
    def close_sql(self):
        '''关闭数据库操作'''
        self.sql_file.close()  # 关闭文件
        self.mysql_connect.close()  # 关闭数据库连接
        self.mysql_cursor.close()  # 关闭游标

    def code(self):
        '''验证码处理'''
        randnum = random.randint(100000, 999999)
        print('----生成的随机数为:', randnum)
        input_num = input(u"请输入验证码")
        print('----输入验证码为:', input_num)
        if input_num == randnum:
            print('随机数正确,登录成功')
        elif input_num == 123456:
            print('输入正确,登录成功')
        else:
            print('登录失败')


    def drag_release(self, source, target):
        '''
        拖拽
        :param source: 
        :param target: 
        :return: 
        '''
        ele_begin = self.locate_element(source)
        ele_end = self.locate_element(target)
        ac = ActionChains(self.driver)
        ac.click_and_hold(ele_begin)
        ac.move_to_element(ele_end)
        ac.release(ele_end).perform()

    def drag_drop(self, selector_source, selector_target):
        '''
        将元素从起点拖到终点后释放
        :param selector_surce: 起点
        :param selector_target: 终点
        :return: 
        '''
        ele_begin = self.locate_element(selector_source)
        ele_end = self.locate_element(selector_target)
        ac = ActionChains(self.driver)
        ac.drag_and_drop(ele_begin, ele_end).perform()

    def count_elements(self, selector):
        '''元素计数'''
        eles = self.locate_elements(selector)

        return len(eles)

    def execute_JavaScript(self, script):
        '''执行JavaScript脚本
        Execute JavaScript scripts.
        Usage: 
        driver.JavaScript("windows.scrollTo(200,1000)")
        '''
        self.driver.execute_script(script)

    def element_selected(self, selector):
        '''返回WebElement的选定状态 
        to return the selected status of an WebElement 

        { isEnable()、isDisplayed()和isSelected() 三个布尔类型的函数 
        isEnable()  用于存储input、select等元素的可编辑状态，可以编辑返回true，否则返回false
        isDisplayed()   用于判断某个元素是否存在页面上
        isSelected()    判断某个元素是否被选中
        }
        :param selector: 
        :return: 
        '''
        ele = self.locate_element(selector)
        return ele.is_selected()

    def element_is_displayed(self,selector):
        ele = self.locate_element(selector)
        return ele.is_displayed()


    def get_url(self):
        '''获取当前页面url'''
        return self.driver.current_url

    def get_title(self):
        '''获取当前页面的标题
        Get window title.
        '''
        return self.driver.title

    def get_text(self, selector):
        '''
        获取元素文本
        :return: 
        '''
        ele = self.locate_element(selector)
        return ele.text

    def get_text_list(self, selector):
        '''
        根据selector获取多个元素，取得元素的text列表
        :return:list
        '''
        ele_list = self.locate_elements(selector)
        texts_list = []
        for ele in ele_list:
            texts_list.append(ele.text)
        return texts_list

    def get_attribute(self, selector, attribute):
        '''获取元素属性值
        Get the value of an element attribute.
        '''
        ele = self.locate_element(selector)
        text = ele.get_attribute(attribute)
        return text
    def get_attributes_list(self, selector, attribute):
        '''获取一组元素属性值列表
        Get the value of an element attribute.
        '''
        ele_list = self.locate_elements(selector)
        attribute_list = []
        for ele in ele_list:
            attribute_list.append(ele.get_attribute(attribute))
        return attribute_list

    def get_attribute_enter(self, selector, attribute):
        '''获取元素属性后点击ENTER'''
        ele = self.locate_element(selector)
        ele.send_keys(Keys.ENTER)
        return ele.get_attribute(attribute)

    def get_display(self, selector):
        '''获取要显示的元素，返回结果真或假。
        Get the element to display,the return result is ture or false.
        '''
        ele = self.locate_element(selector)
        return ele.is_displayed()

    def get_names(self, selector):
        names = self.driver.find_elements_by_name(selector)
        return names

    def get_screenshot(self, path):
        '''截图 .png 格式，显示截图时间'''
        self.driver.get_screenshot_as_file(
            '%s\\%s.png' % (path,time.strftime("%Y.%m.%d-%H %M %S",time.localtime())))

    def get_csv_data(self, file_path, charset):
        '''获取csv文件'''
        self.csv_file = open(file_path, mode="r", encoding=charset)
        '''中文数据出现乱码处理方法:settings-editor-file encodings,设置UTF-8'''
        '''读取文件内容'''
        csv_data = csv.reader(self.csv_file)
        return csv_data

    def get_sql_data(self,file_path,file_reservation,character_set,address,username,password,database,port,charset):
        '''
        读取 xx.sql 文件,操作mysql数据库
        :param file_path: .sql文件路径
        :param file_reservation: .sql文件读取权限
        :param character_set: .sql文件字符集
        :param address: mysql数据库所在机器的地址
        :param username: mysql数据库登录用户名
        :param password: mysql数据库登录密码
        :param database: 数据库
        :param port: 端口
        :param charset: mysql数据库字符集
        :return: sql语句操作的数据
        '''
        self.sql_file = open(file_path,     #xx.sql文件路径
                        mode=file_reservation,           #xx.sql文件读取权限
                        encoding=character_set)      #xx.sql文件字符集（character_set）
        sql_scripts = self.sql_file.read()  # 仅读到sql语句，还没读取到数据
        '''连接数据库,要先在cmd模式下用命令导入pymsql包，pip install pymysql'''
        self.mysql_connect = pymysql.connect(
            host=address,  # 数据库所在机器的地址
            user=username,      #数据库登录用户名
            passwd=password,    #数据库登录密码（为空时可写：None）
            db=database,        # 数据库
            port=port,          # 端口（不要加引号）
            charset=charset)       # 数据库字符集（防止读取数据库数据出现乱码）
        '''创建游标并读取数据库数据'''
        self.mysql_cursor = self.mysql_connect.cursor()  # 创建游标，逐行读取数据
        self.mysql_cursor.execute(sql_scripts)  # 执行sql语句
        mysql_data = self.mysql_cursor.fetchall()  # 接收全部的返回结果行
        return mysql_data

    def get_sql_server_data(self,file_path,file_reservation,character_set,address,username,password,db,charset):
        '''
        读取 xx.sql 文件,操作sql_server数据库
        :param file_path: .sql文件路径
        :param file_reservation: .sql文件读取权限
        :param character_set: .sql文件字符集
        :param address: mssql数据库所在机器的地址
        :param username: mssql数据库登录用户名
        :param password: mssql数据库登录密码
        :param database: 数据库
        :param port: 端口
        :param charset: mssql数据库字符集
        :return: sql语句操作的数据
        '''
        self.sql_file = open(self.PATH + file_path,     #xx.sql文件路径
                        mode=file_reservation,           #xx.sql文件读取权限
                        encoding=character_set
                             )      #xx.sql文件字符集（character_set）
        sql_scripts = self.sql_file.read()  # 仅读到sql语句，还没读取到数据
        '''连接数据库,要先在cmd模式下用命令导入pymsql包，pip install pymysql'''
        self.sql_server_connect = pymssql.connect(
            host=address,  # 数据库所在机器的地址
            user=username,      #数据库登录用户名
            password=password,    #数据库登录密码（为空时可写：None）
            database=db,        # 数据库
            # port=port,          # 端口（不要加引号）  暂时不用
            charset=charset)       # 数据库字符集（防止读取数据库数据出现乱码）
        '''创建游标并读取数据库数据'''
        self.mssql_cursor = self.sql_server_connect.cursor()  # 创建游标，逐行读取数据
        self.mssql_cursor.execute(sql_scripts)  # 执行sql语句
        mssql_data = self.mssql_cursor.fetchall()  # 接收全部的返回结果行
        return mssql_data

    def is_element_exist(self,selector):
        '''
        判断一个元素是否存在
        :return:
        '''
        s = self.locate_element(selector)
        if len(s) == 0:
            print("元素未找到：%s" % selector)
            return False
        elif len(s) == 1:
            return True
        else:
            print("找到%s个元素：%s" % (len(s),selector))

    def isElementExist(self,selector):
        '''
        元素不存在，捕获异常
        :param selector:
        :return:
        '''
        try:
            self.locate_element(selector)
            return True
        except:
            return False

    def implicitly_wait(self, second):
        '''智能等待元素'''
        self.driver.implicitly_wait(second)

    def maximize_window(self):
        '''窗口最大化'''
        self.driver.maximize_window()

    def mouseover_to(self, selector):
        '''将鼠标指针移动到选择器 
        To move mouse pointer to selector
        '''
        ele = self.locate_element(selector)
        ac = ActionChains(self.driver)
        ac.move_to_element(ele).perform()

    def navigate(self, url):
        '''
        打开（url）网页
        :param url: 
        :return: 
        '''
        self.driver.get(url)

    def open_new_window(self, selector):
        '''进入新窗口页面
        Open the new window and switch the handle to the newly opened window.
        '''
        original_handle = self.driver.current_window_handle
        ele = self.locate_element(selector)
        ele.click()
        all_handles = self.driver.window_handles
        for nowly_handle in all_handles:
            if nowly_handle != original_handle:
                self.driver.switch_to.window(nowly_handle)

    def open_old_window(self, current_handle):
        '''进入旧窗口页面'''
        all_handles = self.driver.window_handles
        for handle in all_handles:
            if handle == current_handle:
                self.driver.switch_to.window(handle)

    # '''打开window窗体'''
    # def open_tkinter(self,path,quote,case_text):
    #     '''
    #     lable
    #     :param file:
    #     :param text:
    #     :return:
    #     '''
    #     root = Tk()
    #     try:
    #         text1 = Text(root, height=40, width=70)
    #         photo = PhotoImage(file=self.PATH+path)
    #         text1.insert(END, '\n')
    #         text1.image_create(END, image=photo)
    #         text1.pack(side=LEFT)
    #         text2 = Text(root, height=40, width=70)
    #         scroll = Scrollbar(root, command=text2.yview)
    #         text2.configure(yscrollcommand=scroll.set)
    #         text2.tag_configure('bold_italics', font=('Arial', 13,))
    #         text2.tag_configure('big', font=('Verdana', 20, 'bold'))
    #         text2.tag_configure('color', foreground='#476042',
    #                             font=('Tempus Sans ITC', 18, 'bold'))
    #         text2.tag_bind('follow', '<1>', lambda e, t=text2: t.insert(END, "Not now, maybe later!"))
    #         text2.insert(END, '\n web端自动化测试 \n', 'big')
    #         # quote = """
    #         #   -----------模拟用户登陆测试-----------
    #         # """
    #         text2.insert(END, quote, 'color')
    #         text2.insert(END, '\n   ' + case_text + '\n', 'bold_italics')
    #         text2.pack(side=LEFT)
    #         scroll.pack(side=RIGHT, fill=Y)
    #
    #         # 显示倒计时Label
    #         lbTime = tkinter.Label(root, fg='red', anchor='w')
    #         lbTime.place(x=495,y=500,width=495)
    #         # lbTime.pack()  # 自动调整布局
    #         def autoClose():
    #             for i in range(10):
    #                 lbTime['text'] = '距离窗口关闭还有{}秒'.format(10 - i)
    #                 time.sleep(1)
    #             root.destroy()
    #             root.quit()
    #         t = threading.Thread(target=autoClose)
    #         t.setDaemon(True)
    #         t.start()
    #         root.mainloop()
    #         stop_thread(t)
    #         t.join()
    #
    #
    #     except:
    #         print('未能找到.gif图片路径')
    #
    # def open_tk(self,path,case_text):
    #     root = Tk()
    #     try:
    #         photo = PhotoImage(file=self.PATH+path)  # 声明一下图片
    #         print(self.PATH)
    #         theLabel = Label(root,
    #                          text=case_text,  # 载入文本
    #                          justify=CENTER,  # 声明文本的位置
    #                          image=photo,  # 载入图片
    #                          compound=CENTER,  # 声明图片的位置
    #                          font=("楷体", 20),  # 声明文本字体
    #                          fg="green"  # 声明文本颜色 .
    #                          )
    #         theLabel.pack()  # 自动调整布局
    #         #显示倒计时Label
    #         lbTime = tkinter.Label(root, fg='red', anchor='w')
    #         lbTime.place(x=380, y=260, width=250)
    #         L = threading.Lock()
    #         def autoClose():
    #             L.acquire()
    #             for i in range(10):
    #                 lbTime['text'] = '距离窗口关闭还有{}秒'.format(10 - i)
    #                 time.sleep(1)
    #             root.destroy()
    #             root.quit()
    #             L.release()
    #
    #         tt = threading.Thread(target=autoClose)
    #         tt.setDaemon(True)
    #         tt.start()
    #         root.mainloop()
    #         stop_thread(tt)
    #         tt.join()
    #     except:
    #         print('未能找到.gif图片路径')
    #
    # def open_turtle(self,text,x,y):
    #     try:
    #         turtle.setup(width=0.3, height=0.3)
    #         turtle.screensize(300, 300, "deepskyblue")
    #
    #         turtle.color("white")
    #         turtle.speed(1)
    #         turtle.penup()
    #         time.sleep(0.5)
    #         turtle.goto(x, y)
    #         turtle.speed(1)
    #         turtle.pendown()
    #         turtle.write(text, font=("Times", 18, "bold"))
    #         time.sleep(1)
    #         turtle.color("red")
    #         turtle.speed(1)
    #         turtle.penup()
    #         time.sleep(0.5)
    #         turtle.goto(-80, -80)
    #         turtle.speed(1)
    #         turtle.pendown()
    #         turtle.write("等待系统运行 . . .", font=("Times", 16, "bold"))
    #         # 隐藏箭头
    #         turtle.hideturtle()
    #
    #         # try:
    #         #     # 获取pid并杀死进程
    #         #     pid = os.getpid()
    #         #     print('已杀死pid为%s的进程,　返回值是:%s' % (pid, pid))
    #         #     os.kill(pid, signal.SIGILL)
    #         # except:
    #         #     print('杀死进程失败')
    #         # 暂停界面，使得用户能够看见展示的图形
    #         # turtle.done()
    #     except:
    #         print('有主线程未结束')

    def quit_browser(self):
        '''退出浏览器'''
        self.driver.quit()

    def remove_cookie(self, name):
        '''
        移除指定 name 的 cookie
        :param name: 
        :return: 
        '''
        # 检查 cookie是否存在，存在就移除
        ancient_cookie_value = self.driver.get_cookie(name)
        if ancient_cookie_value is not None:
            self.driver.delete_cookie(name)

    def refresh(self, url=None):
        '''
        刷新页面
        如果 url 是空值，刷新当前页面，否则(获取)刷新制定页面
        :param url: 默认是为空
        :return: 
        '''
        if url is None:
            self.driver.refresh()
        else:
            self.driver.get(url)


    def scrolldown(self):
        '''滚动到最下方'''
        self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    def scrolldown_uniform(self):
        '''匀速滚到最下方'''
        js = "javascript:window.scrollTo(0,document.body.scrollHeight-2500);"
        # js = "var scrollH=document.body.scrollHeight;var scrollC = 0;var setT=setInterval(function(){scrollC  += 10;if(scrollC >= scrollH){scrollC = scrollH;clearInterval(setT)},10);window.scrollTo(0, scrollC);"

        self.driver.execute_script(js)
    def scrolltop(self):
        '''滚动到最上方'''
        self.driver.execute_script("window.scrollTo(document.body.scrollWidth, 0);")

    def scrolltop_uniform(self):
        '''缓慢滚动到顶部'''
        js = "window.scrollTo(document.body.scrollWidth, 0);"
        self.driver.execute_script(js)
    def csrollto(self,num1,num2):
        '''滚动到指定位置'''
        self.driver.execute_script("window.scrollTo(%d, %d);" % (num1,num2))

    def sleep(self, seconds):
        '''
        休眠等待
        :param seconds:
        :return:
        '''
        time.sleep(seconds)


    def switch_to_frame(self, selector):
        '''进入iframe'''
        ele = self.locate_element(selector)
        self.driver.switch_to.frame(ele)

    def switch_to_parent_frame(self):
        '''退出到上一层iframe'''
        self.driver.switch_to.parent_frame()

    def switch_default_frame(self):
        '''退出到最外层iframe'''
        self.driver.switch_to.default_content()

    def switch_to_window_by_title(self, title):
        '''
        通过标题转换窗口
        '''
        for handle in self.driver.window_handles:
            self.driver.switch_to.window(handle)
            if self.driver.title == title:
                break
            self.driver.switch_to.default_content()
    def switch_to_current_window(self):
        '''
        转换到当前窗口
        :return:
        '''
        windows = self.driver.window_handles
        self.driver.switch_to_window(windows[-1])

    def switch_to_window_by_name(self, window_name):
        '''切换窗口'''
        warnings.warn("use driver.switch_to.window instead", DeprecationWarning)
        self.driver.switch_to.window(window_name)

    def switch_to_window_by_handle(self,number):
        '''通过句柄转换窗口'''
        self.driver.switch_to_window(self.driver.window_handles[number])

    def switch_to_default(self):
        '''切换到默认窗口'''
        self.driver.switch_to_default_content()

    def select_by_index(self, selector, num):
        '''以index方式选择下拉框选项'''
        ele = self.locate_element(selector)
        select = Select(ele)
        select.select_by_index(num)

    def select_by_value(self, selector, value):
        '''以value方式选择下拉框选项'''
        ele = self.locate_element(selector)
        select = Select(ele)
        select.select_by_value(value)

    def select_by_visible_text(self, selector, text):
        '''以text方式选择下拉框选项'''
        ele = self.locate_element(selector)
        select = Select(ele)
        select.select_by_visible_text(text)

    def type(self, selector, text):
        '''输入内容'''
        ele = self.locate_element(selector)  # 元素定位
        ele.send_keys(text)

    def type_clear(self, selector, text):
        '''清空后输入内容'''
        ele = self.locate_element(selector)  # 元素定位
        ele.clear()
        ele.send_keys(text)

    def type_enter(self, selector, text):
        """输入内容,点击ENTER"""
        ele = self.locate_element(selector)  # 元素定位
        ele.send_keys(text)
        ele.send_keys(Keys.ENTER)
		
    '''非input标签上传多个（单个）文件'''
    def upload_file_win32gui(self, selector, *file):
        '''
        非input标签上传多个（单个）文件
        :param selector:
        :param file: 自定义传参个数
        :return:
        '''
        # 定位上传文件元素实现点击
        upload = self.locate_element(selector)
        upload.click()
        time.sleep(2)


        dialog = win32gui.FindWindow('#32770', None)  # 对话框
        ComboBoxEx32 = win32gui.FindWindowEx(dialog, 0, 'ComboBoxEx32', None)  # 下拉列表框
        ComboBox = win32gui.FindWindowEx(ComboBoxEx32, 0, 'ComboBox', None)  # 下拉列表框
        Edit = win32gui.FindWindowEx(ComboBox, 0, 'Edit', None)  # 编辑
        button = win32gui.FindWindowEx(dialog, 0, 'Button', None)  # 按钮


        # 跟上面示例的代码是一样的，只是这里传入的参数不同，如果愿意可以写一个上传函数把上传功能封装起来
        win32gui.SendMessage(Edit, win32con.WM_SETTEXT, 0, *file)
        win32gui.SendMessage(dialog, win32con.WM_COMMAND, 1, button)
        # self.driver.quit()

    def window_handles(self):
        '''
        获取所有窗口句柄
        :return: 全部的窗口的句柄
        '''
        return self.driver.window_handles

    def xlsx_csv(self,xlsx_path, sheet,csv_path):
        '''
        获取当前路径的父级路径
        尝试移除旧的CSV文件，如果为空则抛出异常，不影响代码运行
        :param xlsx_path: 读取xlsx表格路径
        :param csv_path: 写入csv表格路径
        :return: csv文件
        '''
        old_file = self.PATH + csv_path
        try:
            os.remove(old_file)
        except Exception as msg:
            print('无旧的CSV文件，正常跳过此步骤', msg)
        wb = load_workbook(self.PATH +xlsx_path)
        ws = wb[sheet]

        for row in ws.iter_rows(min_row=None, max_row=None, min_col=None, max_col=None):
            line = [col.value for col in row]
            with open(self.PATH + csv_path, "a+", newline="", encoding='utf-8') as datacsv:
                csvwriter = csv.writer(datacsv, dialect=("excel"))
                csvwriter.writerow(line)
                datacsv.close()
        wb.close()


